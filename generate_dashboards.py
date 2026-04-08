from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DOWNLOADS = Path.home() / "Downloads"
SOURCE_FILE = DOWNLOADS / "FREQUÊNCIA.xlsx"
OUTPUT_XLSX = ROOT / "dashboard_frequencia_excel.xlsx"
OUTPUT_HTML = ROOT / "dashboard_frequencia.html"
OUTPUT_JSON = ROOT / "dashboard_data.json"
PRIMARY_SHEETS = [
    "GINECOLOGIA ENDOCRINOLÓGICA",
    "DIABETES E OBESIDADE",
    "GERIATRIA",
]
HEADER_FILL = PatternFill("solid", fgColor="16324F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CARD_FILL = PatternFill("solid", fgColor="F8FBFF")
ACCENT_FILL = PatternFill("solid", fgColor="D9EAFD")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7E0EA"),
    right=Side(style="thin", color="D7E0EA"),
    top=Side(style="thin", color="D7E0EA"),
    bottom=Side(style="thin", color="D7E0EA"),
)


def resolve_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    try:
        with path.open("ab"):
            return path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return path.with_name(f"{path.stem}_{timestamp}{path.suffix}")


def find_source_file() -> Path:
    if SOURCE_FILE.exists():
        return SOURCE_FILE

    matches = sorted(DOWNLOADS.glob("FREQU*.xlsx"))
    if matches:
        return matches[0]

    raise FileNotFoundError("Não encontrei a planilha FREQUÊNCIA.xlsx na pasta Downloads.")


def load_data(source_file: Path) -> pd.DataFrame:
    excel = pd.ExcelFile(source_file)
    frames: list[pd.DataFrame] = []

    for sheet_name in excel.sheet_names:
        if sheet_name not in PRIMARY_SHEETS:
            continue

        frame = pd.read_excel(source_file, sheet_name=sheet_name, header=1)
        frame.columns = [str(column).strip() for column in frame.columns]
        frame = frame.dropna(how="all")
        required = ["DATA", "TURNO", "DISCIPLINA", "ALUNO", "PRESENTE", "AUSENTE"]
        if not set(required).issubset(frame.columns):
            continue

        frame = frame[required].copy()
        frame["CURSO"] = sheet_name
        frames.append(frame)

    if not frames:
        raise ValueError("Nenhuma aba válida foi encontrada para consolidar.")

    data = pd.concat(frames, ignore_index=True)
    data["DATA"] = pd.to_datetime(data["DATA"], errors="coerce")
    data["DATA_EXIBICAO"] = data["DATA"].dt.strftime("%d/%m/%Y")
    data["TURNO"] = data["TURNO"].fillna("Não informado").astype(str).str.strip()
    data["DISCIPLINA"] = data["DISCIPLINA"].fillna("Não informada").astype(str).str.strip()
    data["ALUNO"] = data["ALUNO"].fillna("Não informado").astype(str).str.strip()
    data["PRESENTE"] = data["PRESENTE"].fillna(False).astype(bool)
    data["AUSENTE"] = data["AUSENTE"].fillna(False).astype(bool)
    data["STATUS"] = data.apply(
        lambda row: "Presente"
        if row["PRESENTE"]
        else "Ausente"
        if row["AUSENTE"]
        else "Sem marcação",
        axis=1,
    )
    data["MES"] = data["DATA"].dt.strftime("%Y-%m")
    data["AULAS"] = 1
    data = data.sort_values(["DATA", "CURSO", "DISCIPLINA", "ALUNO"]).reset_index(drop=True)
    return data


def summarize_data(data: pd.DataFrame) -> dict[str, pd.DataFrame | dict[str, object]]:
    valid = data[data["STATUS"].isin(["Presente", "Ausente"])].copy()

    total_registros = int(len(valid))
    total_registros_brutos = int(len(data))
    total_registros_invalidos = max(total_registros_brutos - total_registros, 0)
    total_presencas = int(valid["PRESENTE"].sum())
    total_ausencias = int(valid["AUSENTE"].sum())
    frequencia_geral = (total_presencas / total_registros) if total_registros else 0

    course = (
        valid.groupby("CURSO", as_index=False)
        .agg(
            Registros=("AULAS", "sum"),
            Presencas=("PRESENTE", "sum"),
            Ausencias=("AUSENTE", "sum"),
            Alunos=("ALUNO", "nunique"),
            Disciplinas=("DISCIPLINA", "nunique"),
        )
        .sort_values("CURSO")
        .reset_index(drop=True)
    )
    course["Frequencia"] = course["Presencas"] / course["Registros"]

    discipline = (
        valid.groupby(["CURSO", "DISCIPLINA"], as_index=False)
        .agg(
            Registros=("AULAS", "sum"),
            Presencas=("PRESENTE", "sum"),
            Ausencias=("AUSENTE", "sum"),
            Alunos=("ALUNO", "nunique"),
        )
        .sort_values(["Ausencias", "DISCIPLINA"], ascending=[False, True])
        .reset_index(drop=True)
    )
    discipline["Frequencia"] = discipline["Presencas"] / discipline["Registros"]

    student = (
        valid.groupby(["CURSO", "ALUNO"], as_index=False)
        .agg(
            Registros=("AULAS", "sum"),
            Presencas=("PRESENTE", "sum"),
            Ausencias=("AUSENTE", "sum"),
            Disciplinas=("DISCIPLINA", "nunique"),
        )
        .sort_values(["Ausencias", "ALUNO"], ascending=[False, True])
        .reset_index(drop=True)
    )
    student["Frequencia"] = student["Presencas"] / student["Registros"]

    timeline = (
        valid.groupby(["DATA", "DATA_EXIBICAO", "CURSO", "STATUS"], as_index=False)
        .agg(Quantidade=("AULAS", "sum"))
        .sort_values("DATA")
    )

    timeline_global = (
        valid.groupby(["DATA", "DATA_EXIBICAO", "STATUS"], as_index=False)
        .agg(Quantidade=("AULAS", "sum"))
        .sort_values("DATA")
    )
    timeline_pivot = (
        timeline_global.pivot(index=["DATA", "DATA_EXIBICAO"], columns="STATUS", values="Quantidade")
        .fillna(0)
        .reset_index()
    )
    for column in ["Presente", "Ausente"]:
        if column not in timeline_pivot.columns:
            timeline_pivot[column] = 0
    timeline_pivot = timeline_pivot.drop(columns=["DATA"])

    status = (
        valid.groupby("STATUS", as_index=False)
        .agg(Quantidade=("AULAS", "sum"))
        .sort_values("Quantidade", ascending=False)
    )

    top_students = student.head(12).copy()
    top_disciplines = discipline.head(12).copy()

    periodo_inicio = valid["DATA"].min() if total_registros else None
    periodo_fim = valid["DATA"].max() if total_registros else None

    meta = {
        "total_registros": total_registros,
        "total_registros_invalidos": total_registros_invalidos,
        "total_registros_brutos": total_registros_brutos,
        "total_presencas": total_presencas,
        "total_ausencias": total_ausencias,
        "frequencia_geral": frequencia_geral,
        "total_alunos": int(valid["ALUNO"].nunique()),
        "total_disciplinas": int(valid["DISCIPLINA"].nunique()),
        "periodo_inicio": periodo_inicio.strftime("%d/%m/%Y") if periodo_inicio is not None and pd.notna(periodo_inicio) else "",
        "periodo_fim": periodo_fim.strftime("%d/%m/%Y") if periodo_fim is not None and pd.notna(periodo_fim) else "",
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }

    return {
        "meta": meta,
        "course": course,
        "discipline": discipline,
        "student": student,
        "timeline": timeline_pivot,
        "timeline_course": timeline,
        "status": status,
        "top_students": top_students,
        "top_disciplines": top_disciplines,
    }


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)


def add_excel_table(ws, display_name: str) -> None:
    last_row = ws.max_row
    last_col = ws.max_column
    if last_row < 2 or last_col < 1:
        return
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_sheet(ws, dataframe: pd.DataFrame, title: str, as_table: bool = True) -> None:
    ws.title = title
    ws.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False):
        ws.append(list(row))
    style_header(ws[1])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
    if as_table:
        safe_name = "".join(char for char in title.title() if char.isalnum())[:20] or "Tabela"
        add_excel_table(ws, safe_name)
    autosize_columns(ws)


def create_excel_dashboard(data: pd.DataFrame, summary: dict[str, pd.DataFrame | dict[str, object]]) -> Path:
    wb = Workbook()
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    base_columns = ["DATA_EXIBICAO", "CURSO", "TURNO", "DISCIPLINA", "ALUNO", "STATUS"]
    base = data[base_columns].copy()
    write_sheet(wb.create_sheet(), base, "Base Consolidada")
    write_sheet(wb.create_sheet(), summary["course"], "Resumo Cursos")
    write_sheet(wb.create_sheet(), summary["discipline"], "Resumo Disciplinas")
    write_sheet(wb.create_sheet(), summary["student"], "Resumo Alunos")
    write_sheet(wb.create_sheet(), summary["timeline"], "Resumo Diário")
    ws_controls = wb.create_sheet("Controles")

    ws_dashboard["A1"] = "Dashboard de Frequência"
    ws_dashboard["A2"] = "Planilha consolidada com gráficos e indicadores"
    ws_dashboard["A1"].font = Font(size=20, bold=True, color="16324F")
    ws_dashboard["A2"].font = Font(size=11, color="5B6572")
    ws_dashboard["A1"].alignment = Alignment(horizontal="left")

    meta = summary["meta"]
    cards = [
        ("Registros válidos", meta["total_registros"]),
        ("Presenças", meta["total_presencas"]),
        ("Ausências", meta["total_ausencias"]),
        ("Frequência geral", f"{meta['frequencia_geral']:.1%}"),
        ("Alunos", meta["total_alunos"]),
        ("Disciplinas", meta["total_disciplinas"]),
        ("Início", meta["periodo_inicio"]),
        ("Fim", meta["periodo_fim"]),
    ]

    card_fill = PatternFill("solid", fgColor="EAF3FF")
    border_fill = PatternFill("solid", fgColor="DCEBFA")
    positions = ["A", "C", "E", "G", "A", "C", "E", "G"]
    rows = [4, 4, 4, 4, 7, 7, 7, 7]
    for (label, value), column, row in zip(cards, positions, rows):
        ws_dashboard[f"{column}{row}"] = label
        ws_dashboard[f"{column}{row+1}"] = value
        for target in [f"{column}{row}", f"{column}{row+1}"]:
            ws_dashboard[target].fill = card_fill if target.endswith(str(row + 1)) else border_fill
            ws_dashboard[target].alignment = Alignment(horizontal="center", vertical="center")
            ws_dashboard[target].border = THIN_BORDER
        ws_dashboard[f"{column}{row}"].font = Font(size=11, bold=True, color="16324F")
        ws_dashboard[f"{column}{row+1}"].font = Font(size=16, bold=True, color="0F172A")

    ws_dashboard["A10"] = "Gerado em"
    ws_dashboard["B10"] = meta["gerado_em"]
    ws_dashboard["A10"].font = Font(bold=True)
    ws_dashboard["B10"].fill = CARD_FILL
    ws_dashboard["B10"].border = THIN_BORDER

    course_sheet = wb["Resumo Cursos"]
    timeline_sheet = wb["Resumo Diário"]

    chart_courses = BarChart()
    chart_courses.title = "Frequência por Curso"
    chart_courses.y_axis.title = "Frequência"
    chart_courses.x_axis.title = "Curso"
    chart_courses.height = 7
    chart_courses.width = 12
    data_ref = Reference(course_sheet, min_col=6, min_row=1, max_row=1 + len(summary["course"]))
    categories_ref = Reference(course_sheet, min_col=1, min_row=2, max_row=1 + len(summary["course"]))
    chart_courses.add_data(data_ref, titles_from_data=True)
    chart_courses.set_categories(categories_ref)
    chart_courses.style = 10
    ws_dashboard.add_chart(chart_courses, "A12")

    chart_absences = BarChart()
    chart_absences.title = "Ausências por Curso"
    chart_absences.y_axis.title = "Ausências"
    chart_absences.height = 7
    chart_absences.width = 12
    abs_ref = Reference(course_sheet, min_col=4, min_row=1, max_row=1 + len(summary["course"]))
    chart_absences.add_data(abs_ref, titles_from_data=True)
    chart_absences.set_categories(categories_ref)
    chart_absences.style = 11
    ws_dashboard.add_chart(chart_absences, "N12")

    chart_status = PieChart()
    chart_status.title = "Distribuição de Status"
    status_sheet = wb.create_sheet("Resumo Status")
    write_sheet(status_sheet, summary["status"], "Resumo Status")
    pie_labels = Reference(status_sheet, min_col=1, min_row=2, max_row=1 + len(summary["status"]))
    pie_data = Reference(status_sheet, min_col=2, min_row=1, max_row=1 + len(summary["status"]))
    chart_status.add_data(pie_data, titles_from_data=True)
    chart_status.set_categories(pie_labels)
    chart_status.height = 7
    chart_status.width = 10
    chart_status.style = 10
    ws_dashboard.add_chart(chart_status, "A28")

    chart_timeline = LineChart()
    chart_timeline.title = "Evolução Diária"
    chart_timeline.y_axis.title = "Quantidade"
    chart_timeline.height = 7
    chart_timeline.width = 14
    time_data = Reference(timeline_sheet, min_col=2, min_row=1, max_col=3, max_row=1 + len(summary["timeline"]))
    time_categories = Reference(timeline_sheet, min_col=1, min_row=2, max_row=1 + len(summary["timeline"]))
    chart_timeline.add_data(time_data, titles_from_data=True)
    chart_timeline.set_categories(time_categories)
    chart_timeline.style = 13
    ws_dashboard.add_chart(chart_timeline, "N28")

    ws_dashboard.column_dimensions["A"].width = 16
    ws_dashboard.column_dimensions["B"].width = 16
    ws_dashboard.column_dimensions["C"].width = 16
    ws_dashboard.column_dimensions["D"].width = 16
    ws_dashboard.column_dimensions["E"].width = 16
    ws_dashboard.column_dimensions["F"].width = 16
    ws_dashboard.column_dimensions["G"].width = 16
    ws_dashboard.column_dimensions["H"].width = 16
    ws_dashboard.sheet_view.showGridLines = False

    # Controles e filtros práticos
    ws_controls["A1"] = "Controles de Filtro"
    ws_controls["A2"] = "Use as listas abaixo para selecionar rapidamente valores e depois aplique o autofiltro nas abas de base e resumos."
    ws_controls["A1"].font = Font(size=18, bold=True, color="16324F")
    ws_controls["A2"].font = Font(size=10, color="5B6572")
    ws_controls["A4"] = "Curso"
    ws_controls["B4"] = "Disciplina"
    ws_controls["C4"] = "Aluno"
    ws_controls["D4"] = "Turno"
    style_header(ws_controls[4])

    unique_courses = sorted(data["CURSO"].dropna().unique())
    unique_disciplines = sorted(data["DISCIPLINA"].dropna().unique())
    unique_students = sorted(data["ALUNO"].dropna().unique())
    unique_shifts = sorted(data["TURNO"].dropna().unique())

    ws_controls["F1"] = "Listas de apoio"
    ws_controls["F1"].font = Font(bold=True, color="16324F")
    ws_controls["F2"] = "Cursos"
    ws_controls["G2"] = "Disciplinas"
    ws_controls["H2"] = "Alunos"
    ws_controls["I2"] = "Turnos"
    style_header(ws_controls[2])

    max_len = max(len(unique_courses), len(unique_disciplines), len(unique_students), len(unique_shifts))
    for index in range(max_len):
        row = index + 3
        if index < len(unique_courses):
            ws_controls[f"F{row}"] = unique_courses[index]
        if index < len(unique_disciplines):
            ws_controls[f"G{row}"] = unique_disciplines[index]
        if index < len(unique_students):
            ws_controls[f"H{row}"] = unique_students[index]
        if index < len(unique_shifts):
            ws_controls[f"I{row}"] = unique_shifts[index]

    ws_controls["F3"] = "Todos"
    ws_controls["G3"] = "Todas"
    ws_controls["H3"] = "Todos"
    ws_controls["I3"] = "Todos"

    validations = [
        ("A5", f"'Controles'!$F$3:$F${len(unique_courses) + 3}"),
        ("B5", f"'Controles'!$G$3:$G${len(unique_disciplines) + 3}"),
        ("C5", f"'Controles'!$H$3:$H${len(unique_students) + 3}"),
        ("D5", f"'Controles'!$I$3:$I${len(unique_shifts) + 3}"),
    ]
    for cell_ref, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.prompt = "Selecione um valor"
        validation.error = "Escolha um item da lista."
        ws_controls.add_data_validation(validation)
        validation.add(ws_controls[cell_ref])
        ws_controls[cell_ref].fill = CARD_FILL
        ws_controls[cell_ref].border = THIN_BORDER

    ws_controls["A7"] = "Dica"
    ws_controls["B7"] = "Nas abas com tabela, use a seta do cabeçalho para filtrar por curso, disciplina, aluno, turno ou status. As listas acima ajudam na seleção rápida."
    ws_controls["A7"].font = Font(bold=True, color="16324F")
    ws_controls["B7"].font = Font(color="334155")
    ws_controls["B7"].alignment = Alignment(wrap_text=True)
    autosize_columns(ws_controls)
    ws_controls.sheet_view.showGridLines = False

    # Destaques visuais nas abas de resumo
    freq_course_col = summary["course"].columns.get_loc("Frequencia") + 1
    freq_student_col = summary["student"].columns.get_loc("Frequencia") + 1
    freq_discipline_col = summary["discipline"].columns.get_loc("Frequencia") + 1
    abs_student_col = summary["student"].columns.get_loc("Ausencias") + 1
    abs_discipline_col = summary["discipline"].columns.get_loc("Ausencias") + 1

    wb["Resumo Cursos"].conditional_formatting.add(
        f"{get_column_letter(freq_course_col)}2:{get_column_letter(freq_course_col)}{wb['Resumo Cursos'].max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FECACA", mid_type="num", mid_value=0.75, mid_color="FEF3C7", end_type="num", end_value=1, end_color="BBF7D0"),
    )
    wb["Resumo Alunos"].conditional_formatting.add(
        f"{get_column_letter(freq_student_col)}2:{get_column_letter(freq_student_col)}{wb['Resumo Alunos'].max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FECACA", mid_type="num", mid_value=0.75, mid_color="FEF3C7", end_type="num", end_value=1, end_color="BBF7D0"),
    )
    wb["Resumo Disciplinas"].conditional_formatting.add(
        f"{get_column_letter(freq_discipline_col)}2:{get_column_letter(freq_discipline_col)}{wb['Resumo Disciplinas'].max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FECACA", mid_type="num", mid_value=0.75, mid_color="FEF3C7", end_type="num", end_value=1, end_color="BBF7D0"),
    )
    wb["Resumo Alunos"].conditional_formatting.add(
        f"{get_column_letter(abs_student_col)}2:{get_column_letter(abs_student_col)}{wb['Resumo Alunos'].max_row}",
        ColorScaleRule(start_type="min", start_color="E0F2FE", mid_type="percentile", mid_value=50, mid_color="FDE68A", end_type="max", end_color="FCA5A5"),
    )
    wb["Resumo Disciplinas"].conditional_formatting.add(
        f"{get_column_letter(abs_discipline_col)}2:{get_column_letter(abs_discipline_col)}{wb['Resumo Disciplinas'].max_row}",
        ColorScaleRule(start_type="min", start_color="E0F2FE", mid_type="percentile", mid_value=50, mid_color="FDE68A", end_type="max", end_color="FCA5A5"),
    )

    output_path = resolve_output_path(OUTPUT_XLSX)
    wb.save(output_path)
    return output_path


def build_html_payload(data: pd.DataFrame, summary: dict[str, pd.DataFrame | dict[str, object]]) -> dict[str, object]:
    base_records = data[
        ["DATA", "DATA_EXIBICAO", "CURSO", "TURNO", "DISCIPLINA", "ALUNO", "STATUS"]
    ].copy()
    base_records["DATA"] = base_records["DATA"].dt.strftime("%Y-%m-%d")
    base_records["DIA_SEMANA"] = pd.to_datetime(base_records["DATA"]).dt.day_name().map(
        {
            "Monday": "Segunda",
            "Tuesday": "Terça",
            "Wednesday": "Quarta",
            "Thursday": "Quinta",
            "Friday": "Sexta",
            "Saturday": "Sábado",
            "Sunday": "Domingo",
        }
    )
    return {
        "meta": summary["meta"],
        "records": base_records.to_dict(orient="records"),
    }


def create_data_json(data: pd.DataFrame, summary: dict[str, pd.DataFrame | dict[str, object]]) -> Path:
    output_path = resolve_output_path(OUTPUT_JSON)
    payload = build_html_payload(data, summary)
    output_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return output_path


def create_html_dashboard(summary: dict[str, pd.DataFrame | dict[str, object]]) -> Path:
    payload = build_html_payload(summary)
    payload_json = json.dumps(payload, ensure_ascii=False)
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Dashboard de Frequência</title>
  <style>
    :root {{
      --bg: #f4f7fb;
      --card: rgba(255,255,255,0.92);
      --text: #102033;
      --muted: #5f6b7a;
      --blue: #1d4ed8;
      --green: #15803d;
      --red: #b91c1c;
      --gold: #d97706;
      --stroke: rgba(15, 23, 42, 0.08);
      --shadow: 0 18px 40px rgba(15, 23, 42, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "Helvetica Neue", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(59,130,246,0.18), transparent 32%),
        radial-gradient(circle at right center, rgba(245,158,11,0.14), transparent 25%),
        linear-gradient(180deg, #f8fbff 0%, var(--bg) 100%);
    }}
    .wrap {{
      max-width: 1240px;
      margin: 0 auto;
      padding: 32px 20px 48px;
    }}
    .hero {{
      display: flex;
      justify-content: space-between;
      gap: 20px;
      align-items: end;
      margin-bottom: 24px;
    }}
    .hero h1 {{
      margin: 0;
      font-size: clamp(2rem, 4vw, 3.2rem);
      line-height: 1;
      letter-spacing: -0.03em;
    }}
    .hero p {{
      margin: 10px 0 0;
      color: var(--muted);
      max-width: 760px;
      font-size: 1rem;
    }}
    .stamp {{
      background: rgba(255,255,255,0.85);
      border: 1px solid var(--stroke);
      box-shadow: var(--shadow);
      border-radius: 18px;
      padding: 16px 18px;
      min-width: 220px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 16px;
      margin-bottom: 18px;
    }}
    .card, .panel {{
      background: var(--card);
      border: 1px solid var(--stroke);
      box-shadow: var(--shadow);
      border-radius: 22px;
      backdrop-filter: blur(10px);
    }}
    .card {{
      padding: 18px;
    }}
    .label {{
      color: var(--muted);
      font-size: 0.9rem;
      margin-bottom: 12px;
    }}
    .value {{
      font-size: 2rem;
      font-weight: 700;
      letter-spacing: -0.03em;
    }}
    .panel {{
      padding: 20px;
    }}
    .panels {{
      display: grid;
      grid-template-columns: 1.4fr 1fr;
      gap: 16px;
      margin-bottom: 16px;
    }}
    .filters {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 14px;
      margin-bottom: 16px;
    }}
    .filters label {{
      display: block;
      font-size: 0.82rem;
      color: var(--muted);
      margin-bottom: 8px;
      text-transform: uppercase;
      letter-spacing: 0.05em;
    }}
    .filters input, .filters select {{
      width: 100%;
      border: 1px solid rgba(15, 23, 42, 0.1);
      background: rgba(255,255,255,0.9);
      border-radius: 14px;
      padding: 12px 14px;
      color: var(--text);
      font-size: 0.96rem;
      outline: none;
    }}
    .panel h2 {{
      margin: 0 0 18px;
      font-size: 1.1rem;
    }}
    .bars {{
      display: grid;
      gap: 12px;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 220px 1fr 68px;
      gap: 12px;
      align-items: center;
    }}
    .bar-track {{
      height: 14px;
      background: #e7edf6;
      border-radius: 999px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      border-radius: 999px;
      background: linear-gradient(90deg, #2563eb 0%, #38bdf8 100%);
    }}
    .bar-fill.red {{
      background: linear-gradient(90deg, #f97316 0%, #dc2626 100%);
    }}
    .mini-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 0.95rem;
    }}
    th, td {{
      padding: 10px 0;
      border-bottom: 1px solid rgba(15, 23, 42, 0.08);
      text-align: left;
    }}
    th {{
      color: var(--muted);
      font-weight: 600;
      font-size: 0.82rem;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .timeline {{
      height: 290px;
      display: flex;
      align-items: end;
      gap: 8px;
      padding-top: 18px;
    }}
    .timeline-col {{
      flex: 1;
      display: flex;
      flex-direction: column;
      justify-content: end;
      align-items: center;
      gap: 8px;
      min-width: 0;
    }}
    .stack {{
      width: 100%;
      max-width: 42px;
      display: flex;
      flex-direction: column;
      justify-content: end;
      border-radius: 14px 14px 8px 8px;
      overflow: hidden;
      background: #edf2f8;
      min-height: 14px;
    }}
    .present {{
      background: linear-gradient(180deg, #22c55e 0%, #15803d 100%);
    }}
    .absent {{
      background: linear-gradient(180deg, #fb7185 0%, #b91c1c 100%);
    }}
    .date-label {{
      writing-mode: vertical-rl;
      transform: rotate(180deg);
      font-size: 0.72rem;
      color: var(--muted);
      letter-spacing: 0.04em;
    }}
    .foot {{
      margin-top: 18px;
      color: var(--muted);
      font-size: 0.86rem;
      text-align: center;
    }}
    @media (max-width: 980px) {{
      .grid, .panels, .mini-grid {{
        grid-template-columns: 1fr;
      }}
      .filters {{
        grid-template-columns: 1fr;
      }}
      .hero {{
        flex-direction: column;
        align-items: start;
      }}
      .bar-row {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <div>
        <h1>Dashboard de Frequência</h1>
        <p>Visão consolidada das turmas de pós-graduação com foco em presença, ausências, frequência por curso e pontos de atenção por aluno e disciplina.</p>
      </div>
      <div class="stamp">
        <div class="label">Arquivo gerado</div>
        <div class="value" style="font-size:1.2rem;" id="gerado-em"></div>
      </div>
    </section>

    <section class="grid">
      <article class="card"><div class="label">Registros válidos</div><div class="value" id="registros"></div></article>
      <article class="card"><div class="label">Frequência geral</div><div class="value" id="frequencia"></div></article>
      <article class="card"><div class="label">Alunos</div><div class="value" id="alunos"></div></article>
      <article class="card"><div class="label">Disciplinas</div><div class="value" id="disciplinas"></div></article>
    </section>

    <section class="panel" style="margin-bottom:16px;">
      <h2>Filtros rápidos</h2>
      <div class="filters">
        <div>
          <label for="course-filter">Curso</label>
          <select id="course-filter"></select>
        </div>
        <div>
          <label for="discipline-filter">Disciplina</label>
          <input id="discipline-filter" type="text" placeholder="Digite parte do nome" />
        </div>
        <div>
          <label for="student-filter">Aluno</label>
          <input id="student-filter" type="text" placeholder="Buscar aluno" />
        </div>
        <div>
          <label for="min-frequency">Faixa de frequência</label>
          <select id="min-frequency">
            <option value="0">Sem corte</option>
            <option value="below_75">Abaixo de 75%</option>
            <option value="0.75">75%+</option>
            <option value="0.8">80%+</option>
            <option value="0.9">90%+</option>
          </select>
        </div>
      </div>
    </section>

    <section class="panels">
      <article class="panel">
        <h2>Frequência por curso</h2>
        <div class="bars" id="course-bars"></div>
      </article>
      <article class="panel">
        <h2>Resumo executivo</h2>
        <table>
          <tbody>
            <tr><th>Presenças</th><td id="presencas"></td></tr>
            <tr><th>Ausências</th><td id="ausencias"></td></tr>
            <tr><th>Período</th><td id="periodo"></td></tr>
            <tr><th>Cursos</th><td id="cursos-total"></td></tr>
          </tbody>
        </table>
      </article>
    </section>

    <section class="mini-grid">
      <article class="panel">
        <h2>Top alunos com mais ausências</h2>
        <table>
          <thead><tr><th>Aluno</th><th>Curso</th><th>Ausências</th><th>Freq.</th></tr></thead>
          <tbody id="students-table"></tbody>
        </table>
      </article>
      <article class="panel">
        <h2>Disciplinas com mais ausências</h2>
        <table>
          <thead><tr><th>Disciplina</th><th>Curso</th><th>Ausências</th><th>Freq.</th></tr></thead>
          <tbody id="disciplines-table"></tbody>
        </table>
      </article>
    </section>

    <section class="panel" style="margin-top:16px;">
      <h2>Evolução diária de presença e ausência</h2>
      <div class="timeline" id="timeline"></div>
    </section>

    <div class="foot">Abra este arquivo localmente no navegador. Nenhuma instalação é necessária.</div>
  </div>

  <script>
    const data = {payload_json};

    const number = (value) => new Intl.NumberFormat('pt-BR').format(value);
    const percent = (value) => (value * 100).toFixed(1).replace('.', ',') + '%';

    let currentCourse = 'Todos';
    let currentDiscipline = '';
    let currentStudent = '';
    let frequencyMode = '0';

    document.getElementById('gerado-em').textContent = data.meta.gerado_em;
    const courseSelect = document.getElementById('course-filter');
    courseSelect.innerHTML = `<option>Todos</option>${{data.course.map((item) => `<option>${{item.CURSO}}</option>`).join('')}}`;
    courseSelect.addEventListener('change', () => {{
      currentCourse = courseSelect.value;
      render();
    }});
    document.getElementById('discipline-filter').addEventListener('input', (event) => {{
      currentDiscipline = event.target.value.trim().toLowerCase();
      render();
    }});
    document.getElementById('student-filter').addEventListener('input', (event) => {{
      currentStudent = event.target.value.trim().toLowerCase();
      render();
    }});
    document.getElementById('min-frequency').addEventListener('change', (event) => {{
      frequencyMode = event.target.value;
      render();
    }});

    function matchesFrequency(value) {{
      if (frequencyMode === 'below_75') return value < 0.75;
      return value >= Number(frequencyMode);
    }}

    function matchesCourse(item) {{
      return currentCourse === 'Todos' || item.CURSO === currentCourse;
    }}

    function render() {{
      const filteredCourses = data.course.filter((item) => matchesCourse(item) && matchesFrequency(item.Frequencia));
      const filteredStudents = data.students.filter((item) =>
        matchesCourse(item) &&
        matchesFrequency(item.Frequencia) &&
        item.ALUNO.toLowerCase().includes(currentStudent)
      );
      const filteredDisciplines = data.disciplines.filter((item) =>
        matchesCourse(item) &&
        matchesFrequency(item.Frequencia) &&
        item.DISCIPLINA.toLowerCase().includes(currentDiscipline)
      );

      const courseBars = document.getElementById('course-bars');
      courseBars.innerHTML = '';
      filteredCourses.forEach((item) => {{
        const row = document.createElement('div');
        row.className = 'bar-row';
        row.innerHTML = `
          <div><strong>${{item.CURSO}}</strong><div class="label">${{number(item.Alunos)}} alunos</div></div>
          <div class="bar-track"><div class="bar-fill" style="width:${{Math.max(item.Frequencia * 100, 4)}}%"></div></div>
          <div style="text-align:right;font-weight:700;">${{percent(item.Frequencia)}}</div>
        `;
        courseBars.appendChild(row);
      }});

      const studentsTable = document.getElementById('students-table');
      studentsTable.innerHTML = '';
      filteredStudents.slice(0, 12).forEach((item) => {{
        const row = document.createElement('tr');
        row.innerHTML = `<td>${{item.ALUNO}}</td><td>${{item.CURSO}}</td><td>${{item.Ausencias}}</td><td>${{percent(item.Frequencia)}}</td>`;
        studentsTable.appendChild(row);
      }});

      const disciplinesTable = document.getElementById('disciplines-table');
      disciplinesTable.innerHTML = '';
      filteredDisciplines.slice(0, 12).forEach((item) => {{
        const row = document.createElement('tr');
        row.innerHTML = `<td>${{item.DISCIPLINA}}</td><td>${{item.CURSO}}</td><td>${{item.Ausencias}}</td><td>${{percent(item.Frequencia)}}</td>`;
        disciplinesTable.appendChild(row);
      }});

      const totalRegistros = filteredCourses.reduce((acc, item) => acc + item.Registros, 0);
      const totalPresencas = filteredCourses.reduce((acc, item) => acc + item.Presencas, 0);
      const totalAusencias = filteredCourses.reduce((acc, item) => acc + item.Ausencias, 0);
      const totalAlunos = new Set(filteredStudents.map((item) => item.CURSO + '::' + item.ALUNO)).size;
      const totalDisciplinas = new Set(filteredDisciplines.map((item) => item.CURSO + '::' + item.DISCIPLINA)).size;
      const freqGeral = totalRegistros ? totalPresencas / totalRegistros : 0;

      document.getElementById('registros').textContent = number(totalRegistros);
      document.getElementById('frequencia').textContent = percent(freqGeral);
      document.getElementById('alunos').textContent = number(totalAlunos);
      document.getElementById('disciplinas').textContent = number(totalDisciplinas);
      document.getElementById('presencas').textContent = number(totalPresencas);
      document.getElementById('ausencias').textContent = number(totalAusencias);
      document.getElementById('periodo').textContent = `${{data.meta.periodo_inicio}} a ${{data.meta.periodo_fim}}`;
      document.getElementById('cursos-total').textContent = number(filteredCourses.length);

      const timeline = document.getElementById('timeline');
      timeline.innerHTML = '';
      const timelineRows = currentCourse === 'Todos'
        ? data.timeline
        : (() => {{
            const grouped = new Map();
            data.timelineByCourse
              .filter((item) => item.CURSO === currentCourse)
              .forEach((item) => {{
                const key = item.DATA_EXIBICAO;
                if (!grouped.has(key)) grouped.set(key, {{ DATA_EXIBICAO: key, Presente: 0, Ausente: 0 }});
                const row = grouped.get(key);
                row[item.STATUS] = item.Quantidade;
              }});
            return Array.from(grouped.values());
          }})();
      const maxTotal = Math.max(...timelineRows.map((item) => (item.Presente || 0) + (item.Ausente || 0)), 1);
      timelineRows.forEach((item) => {{
        const total = (item.Presente || 0) + (item.Ausente || 0);
        const presentHeight = total ? ((item.Presente || 0) / maxTotal) * 220 : 0;
        const absentHeight = total ? ((item.Ausente || 0) / maxTotal) * 220 : 0;
        const col = document.createElement('div');
        col.className = 'timeline-col';
        col.innerHTML = `
          <div class="stack" title="Presenças: ${{item.Presente || 0}} | Ausências: ${{item.Ausente || 0}}">
            <div class="present" style="height:${{presentHeight}}px"></div>
            <div class="absent" style="height:${{absentHeight}}px"></div>
          </div>
          <div class="date-label">${{item.DATA_EXIBICAO}}</div>
        `;
        timeline.appendChild(col);
      }});
    }}

    render();
  </script>
</body>
</html>
"""
    output_path = resolve_output_path(OUTPUT_HTML)
    output_path.write_text(html, encoding="utf-8")
    return output_path


def main() -> None:
    source_file = find_source_file()
    data = load_data(source_file)
    summary = summarize_data(data)
    excel_path = create_excel_dashboard(data, summary)
    html_path = create_html_dashboard(summary)
    print(json.dumps({
        "source": str(source_file),
        "excel": str(excel_path),
        "html": str(html_path),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
